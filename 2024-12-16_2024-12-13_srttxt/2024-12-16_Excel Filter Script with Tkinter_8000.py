import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import os

def process_excel_file():
    # 메인 윈도우 숨기기
    root = tk.Tk()
    root.withdraw()

    # 파일 선택 다이얼로그
    input_file = filedialog.askopenfilename(
        title="엑셀 파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if not input_file:  # 파일 선택 취소시
        return
    
    try:
        # 엑셀 파일 읽기
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        
        # 헤더 찾기 (검색량 열의 인덱스 찾기)
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        search_col_idx = None
        for idx, header in enumerate(header_row, 1):
            if header == '검색량':
                search_col_idx = idx
                break
                
        if search_col_idx is None:
            messagebox.showerror("에러", "'검색량' 열을 찾을 수 없습니다.")
            return
            
        # 새 워크북 생성
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # 헤더 복사
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(1, col, sheet.cell(1, col).value)
        
        # 데이터 필터링 (8000 이상인 행만 복사)
        new_row = 2
        for row in range(2, sheet.max_row + 1):
            search_value = sheet.cell(row, search_col_idx).value
            try:
                # 숫자가 문자열로 되어있을 경우 쉼표 제거
                if isinstance(search_value, str):
                    search_value = int(search_value.replace(',', ''))
                else:
                    search_value = int(search_value)
                    
                if search_value >= 8000:
                    for col in range(1, sheet.max_column + 1):
                        new_sheet.cell(new_row, col, sheet.cell(row, col).value)
                    new_row += 1
            except (ValueError, TypeError):
                continue
        
        # 저장할 파일 경로 생성
        file_dir = os.path.dirname(input_file)
        file_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(file_dir, f"{file_name}_필터링됨.xlsx")
        
        # 필터링된 데이터 저장
        new_wb.save(output_file)
        
        messagebox.showinfo("완료", f"파일이 저장되었습니다:\n{output_file}")
            
    except Exception as e:
        messagebox.showerror("에러", f"처리 중 오류가 발생했습니다:\n{str(e)}")

if __name__ == "__main__":
    process_excel_file()