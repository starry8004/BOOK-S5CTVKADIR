import os
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from openpyxl import load_workbook, Workbook

def filter_competition_rate(input_path, output_path):
    """
    Filters out rows with competition rate >= 3.1 from Excel files.
    
    Parameters:
    input_path (str): Path to the input Excel file.
    output_path (str): Path to save the filtered Excel file.
    """
    try:
        # Load input workbook and select active sheet
        wb = load_workbook(input_path)
        sheet = wb.active
        
        # Create new workbook and select sheet
        new_wb = Workbook()
        new_sheet = new_wb.active
        
        # Copy header row
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(row=1, column=col).value = sheet.cell(row=1, column=col).value
        
        # Find column index for '경쟁률'
        competition_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == '경쟁률':
                competition_col = col
                break
                
        if competition_col is None:
            raise ValueError("'경쟁률' column not found")
        
        # Filter and copy rows where competition rate < 3.1
        new_row = 2  # Start from second row (after header)
        for row in range(2, sheet.max_row + 1):
            competition_value = sheet.cell(row=row, column=competition_col).value
            if competition_value is not None and float(competition_value) < 3.1:
                for col in range(1, sheet.max_column + 1):
                    new_sheet.cell(row=new_row, column=col).value = sheet.cell(row=row, column=col).value
                new_row += 1
        
        # Save the new workbook
        new_wb.save(output_path)
        print(f"Filtered file saved to: {output_path}")
        
    except Exception as e:
        print(f"An error occurred: {e}")

def generate_output_filename(input_filename):
    """
    Generates new filename with competition rate filter information.
    """
    # Extract base parts from the original filename
    base_parts = input_filename.split('_')
    
    # Get current timestamp
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M")
    
    # Keep essential parts and add new filter information
    new_filename = f"셀하 아이템 발굴 EXCEL_{base_parts[2]}_{base_parts[3]}_경쟁률필터_{current_time}.xlsx"
    
    return new_filename

def main():
    # Create a folder dialog for selecting the folder with Excel files
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    folder_path = filedialog.askdirectory(
        title="Select a Folder Containing Excel Files"
    )
    
    if not folder_path:
        print("No folder selected. Exiting.")
        return

    # Iterate over all Excel files in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            input_file = os.path.join(folder_path, filename)
            output_file = os.path.join(folder_path, generate_output_filename(filename))
            
            # Filter competition rate and save as new Excel file
            filter_competition_rate(input_file, output_file)

    print("Processing completed for all files in the folder.")

if __name__ == "__main__":
    main()