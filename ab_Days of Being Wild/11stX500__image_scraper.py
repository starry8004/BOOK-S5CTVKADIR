import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# 파일 선택창 열기
Tk().withdraw()  # Tkinter 창 숨기기
file_name = askopenfilename(filetypes=[("Excel files", "*.xlsx")])

# 엑셀 파일 열기
wb = openpyxl.load_workbook(file_name)
ws = wb.active

# 새로운 엑셀 파일 생성
new_file_name = f"bestseller_with_images_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = "Best Sellers with Images"
new_ws.append(['순서', '링크', '상품명', '할인율', '정상가', '판매가', '배송비', '판매처', '대표 이미지', '상품 정보 이미지'])

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

# 엑셀 데이터 순회
for row in ws.iter_rows(min_row=2, values_only=True):
    index, product_link, product_name, discount, normal_price, sale_price, delivery, seller = row
    
    try:
        # 개별 상품 페이지 접근
        product_response = requests.get(product_link, headers=headers)
        product_soup = BeautifulSoup(product_response.text, 'html.parser')
        
        # 대표 이미지 추출 (첫 번째 img 태그)
        thumbnail_image = product_soup.select_one('img')['src'] if product_soup.select_one('img') else "N/A"
        
        # 상품 정보 이미지 추출 (상세 이미지 찾기)
        info_images = [img['src'] for img in product_soup.select('img') if '상세' in img['alt']]  # alt에 '상세' 포함된 이미지
        info_images_str = ', '.join(info_images) if info_images else "N/A"
        
        # 새로운 엑셀에 데이터 추가
        new_ws.append([index, product_link, product_name, discount, normal_price, sale_price, delivery, seller, thumbnail_image, info_images_str])
        print(f"순서: {index}, 상품명: {product_name}, 대표 이미지: {thumbnail_image}")
        
    except Exception as e:
        print(f"Error: {e}")

    time.sleep(1)  # 요청 간의 대기 시간 설정

# 새로운 엑셀 파일 저장
new_wb.save(new_file_name)
print(f"{new_file_name} 파일로 저장되었습니다.")
