import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
import time
from tkinter import Tk, filedialog

def select_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

# 사용자에게 입력 파일 선택 요청
input_file = select_file()
if not input_file:
    print("파일이 선택되지 않았습니다. 프로그램을 종료합니다.")
    exit()

# 선택된 엑셀 파일 읽기
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# 새로운 열 추가
ws['I1'] = '대표이미지URL'
ws['J1'] = '상세페이지이미지URL'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

# 각 상품 링크에 대해 추가 정보 스크래핑
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    url = row[0].value
    if url and url.startswith('http'):
        try:
            response = requests.get(url, headers=headers)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 대표 이미지 URL 추출 (예시, 실제 웹사이트 구조에 맞게 수정 필요)
            main_images = soup.select('div.img_full img')
            main_image_urls = ','.join([img['src'] for img in main_images if 'src' in img.attrs])
            
            # 상세 페이지 이미지 URL 추출 (예시, 실제 웹사이트 구조에 맞게 수정 필요)
            detail_images = soup.select('div.product_detail img')
            detail_image_urls = ','.join([img['src'] for img in detail_images if 'src' in img.attrs])
            
            # 데이터 추가
            ws.cell(row=row[0].row, column=9, value=main_image_urls)
            ws.cell(row=row[0].row, column=10, value=detail_image_urls)
            
            print(f"처리 완료: {url}")
            
        except Exception as e:
            print(f"Error processing {url}: {e}")
        
        time.sleep(1)  # 요청 간의 대기 시간 설정

# 결과 저장
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = input_file.rsplit('.', 1)[0] + f'_imageurl_{current_time}.xlsx'
wb.save(output_file)
print(f"데이터가 {output_file}에 저장되었습니다.")

print("프로그램이 완료되었습니다.")