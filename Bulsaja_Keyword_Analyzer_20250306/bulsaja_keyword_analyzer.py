import openpyxl
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
import threading

def to_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def process_file(file_path, progress_data):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        progress_data["error"] = f"파일 로드 실패: {e}"
        progress_data["completed"] = True
        return

    # 헤더 읽기
    headers = [cell.value for cell in sheet[1]]
    
    # 필수 컬럼 인덱스 확인
    try:
        idx_recent = headers.index("최근\n30일\n검색량")
        idx_competition = headers.index("네이버\n경쟁강도")
    except ValueError as e:
        progress_data["error"] = f"필수 컬럼이 없습니다: {e}"
        progress_data["completed"] = True
        return

    try:
        idx_rocket = headers.index("쿠팡\n로켓\n+\n그로스\n비율")
    except ValueError:
        idx_rocket = None

    new_headers = headers + ["base_score", "wing_ratio", "score_adjusted"]
    
    data_rows = []
    total_rows = sheet.max_row - 1  # 헤더 제외
    progress_data["phase"] = "processing"
    progress_data["total"] = total_rows
    progress_data["processed"] = 0

    # 각 행 처리 (데이터 처리 단계)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row = list(row)
        recent_val = to_float(row[idx_recent])
        comp_val = to_float(row[idx_competition])
        base_score = recent_val / (comp_val + 1)
        if idx_rocket is not None:
            rocket_val = to_float(row[idx_rocket])
            wing_ratio = 100 - rocket_val
        else:
            wing_ratio = 0.0
        score_adjusted = base_score * (1 + wing_ratio / 100)
        new_row = row + [base_score, wing_ratio, score_adjusted]
        data_rows.append(new_row)
        progress_data["processed"] += 1

    # 보정 점수 기준 내림차순 정렬
    data_rows.sort(key=lambda r: r[-1], reverse=True)

    # 저장 단계 시작
    progress_data["phase"] = "saving"
    total_save = len(data_rows) + 1  # 헤더 포함
    progress_data["total_save"] = total_save
    progress_data["saved"] = 0

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # 헤더 저장
    new_sheet.append(new_headers)
    progress_data["saved"] += 1

    # 데이터 행 저장 (한 행씩 저장하면서 진행 상황 업데이트)
    for r in data_rows:
        new_sheet.append(r)
        progress_data["saved"] += 1

    current_date = datetime.now().strftime("%Y-%m-%d")
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    new_file_name = f"{base_name}_{current_date}_키워드선별결과.xlsx"
    save_path = os.path.join(os.path.dirname(file_path), new_file_name)
    
    try:
        new_wb.save(save_path)
        progress_data["save_path"] = save_path
    except Exception as e:
        progress_data["error"] = f"저장 실패: {e}"
    
    progress_data["completed"] = True

def update_progress(progress_data, progress_bar, progress_label, root):
    phase = progress_data.get("phase", "processing")
    if phase == "processing":
        total = progress_data.get("total", 0)
        processed = progress_data.get("processed", 0)
        progress_bar["maximum"] = total
        progress_bar["value"] = processed
        progress_label.config(text=f"Processing row {processed} of {total}")
    elif phase == "saving":
        total_save = progress_data.get("total_save", 1)
        saved = progress_data.get("saved", 0)
        progress_bar["maximum"] = total_save
        progress_bar["value"] = saved
        progress_label.config(text=f"Saving row {saved} of {total_save}")
    else:
        progress_label.config(text="Initializing...")
    
    if progress_data.get("completed", False):
        if "error" in progress_data:
            messagebox.showerror("Error", progress_data["error"])
        else:
            messagebox.showinfo("Done", f"Analysis result saved to:\n{progress_data.get('save_path', 'Unknown')}")
        root.destroy()
    else:
        root.after(100, update_progress, progress_data, progress_bar, progress_label, root)

# Tkinter GUI 설정
root = tk.Tk()
root.title("Processing...")
root.geometry("400x150")

progress_label = tk.Label(root, text="Starting processing...")
progress_label.pack(pady=20)
progress_bar = Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=10)

# 파일 선택 창
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
if not file_path:
    messagebox.showerror("Error", "No file selected.")
    root.destroy()
    exit()

# 진행 상황 공유 딕셔너리 생성
progress_data = {"processed": 0, "total": 0, "completed": False, "phase": "processing"}

# 별도 스레드에서 파일 처리 실행
thread = threading.Thread(target=process_file, args=(file_path, progress_data))
thread.start()

# 100ms마다 진행 상황 업데이트
root.after(100, update_progress, progress_data, progress_bar, progress_label, root)
root.mainloop()
