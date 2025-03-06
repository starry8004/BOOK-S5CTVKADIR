import openpyxl
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
import threading

def to_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def calculate_keyword_score(search_volume, competition, wing_ratio,
                              type_weight, intent_weight, competitor_weight, platform_weight, wing_weight):
    # 기본 점수: 검색량 / (경쟁 강도 + 1)
    base_score = search_volume / (competition + 1)
    # wing_ratio는 퍼센트로 들어오므로 100으로 나눠 0~1 사이 값으로 변환
    final_score = base_score * (1 + type_weight + intent_weight + competitor_weight + platform_weight + wing_weight * (wing_ratio / 100))
    return final_score

def process_file(file_path, progress_data, weights, top_N=100):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        progress_data["error"] = f"파일 로드 실패: {e}"
        progress_data["completed"] = True
        return

    headers = [cell.value for cell in sheet[1]]
    try:
        idx_recent = headers.index("최근\n30일\n검색량")
        idx_competition = headers.index("네이버\n경쟁강도")
    except ValueError as e:
        progress_data["error"] = f"필수 컬럼이 없습니다: {e}"
        progress_data["completed"] = True
        return

    # "쿠팡\n로켓\n+\n그로스\n비율" 컬럼이 있으면 해당 인덱스를, 없으면 None
    try:
        idx_rocket = headers.index("쿠팡\n로켓\n+\n그로스\n비율")
    except ValueError:
        idx_rocket = None

    # 새 헤더에 최종 점수(final_score) 추가
    new_headers = headers + ["final_score"]
    data_rows = []
    total_rows = sheet.max_row - 1  # 헤더 제외
    progress_data["total"] = total_rows
    progress_data["processed"] = 0

    # 각 행을 처리하여 최종 점수 계산
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row = list(row)
        search_volume = to_float(row[idx_recent])
        competition = to_float(row[idx_competition])
        # wing_ratio 계산: "쿠팡\n로켓\n+\n그로스\n비율" 값이 있으면 wing_ratio = 100 - 해당 값, 없으면 0
        if idx_rocket is not None:
            rocket_value = to_float(row[idx_rocket])
            wing_ratio = 100 - rocket_value
        else:
            wing_ratio = 0.0

        final_score = calculate_keyword_score(
            search_volume, competition, wing_ratio,
            weights.get("type_weight", 0),
            weights.get("intent_weight", 0),
            weights.get("competitor_weight", 0),
            weights.get("platform_weight", 0),
            weights.get("wing_weight", 0)
        )
        new_row = row + [final_score]
        data_rows.append(new_row)
        progress_data["processed"] += 1

    # 내림차순 정렬하여 상위 키워드 추출
    data_rows.sort(key=lambda r: r[-1], reverse=True)
    top_keywords = data_rows[:top_N]

    # 새 워크북 생성 후 데이터 기록
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.append(new_headers)
    for r in top_keywords:
        new_sheet.append(r)

    current_date = datetime.now().strftime("%Y-%m-%d")
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    new_file_name = f"{base_name}_{current_date}_키워드선별결과.xlsx"
    save_path = os.path.join(os.path.dirname(file_path), new_file_name)
    try:
        new_wb.save(save_path)
        progress_data["save_path"] = save_path
    except Exception as e:
        progress_data["error"] = f"저장 실패: {e}"
    progress_data["completed"] = True

def update_progress(progress_data, progress_bar, progress_label, root):
    total = progress_data.get("total", 0)
    processed = progress_data.get("processed", 0)
    progress_bar["maximum"] = total
    progress_bar["value"] = processed
    progress_label.config(text=f"Processing row {processed} of {total}")

    if progress_data.get("completed", False):
        if "error" in progress_data:
            messagebox.showerror("Error", progress_data["error"])
        else:
            messagebox.showinfo("Done", f"Analysis result saved to:\n{progress_data.get('save_path', 'Unknown')}")
        root.destroy()
    else:
        root.after(100, update_progress, progress_data, progress_bar, progress_label, root)

# Tkinter GUI 설정
root = tk.Tk()
root.title("Keyword Analysis Processing")
root.geometry("400x150")

progress_label = tk.Label(root, text="Starting processing...")
progress_label.pack(pady=20)
progress_bar = Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=10)

# 파일 선택 창
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
if not file_path:
    messagebox.showerror("Error", "No file selected.")
    root.destroy()
    exit()

# 가중치 설정 – 여기서 wing_weight를 추가하여 wing_ratio의 영향력을 반영합니다.
weights = {
    "type_weight": 0.1,         # 키워드 유형 관련 가중치
    "intent_weight": 0.05,      # 고객 의도 반영 가중치
    "competitor_weight": -0.05, # 경쟁사 분석에 따른 가중치 (경쟁 치열하면 마이너스)
    "platform_weight": 0.1,     # 플랫폼 선호도 가중치
    "wing_weight": 0.2          # wing_ratio가 높을수록 최종 점수를 높게 반영 (예: 0.2의 가중치)
}

progress_data = {"processed": 0, "total": 0, "completed": False}
thread = threading.Thread(target=process_file, args=(file_path, progress_data, weights))
thread.start()

root.after(100, update_progress, progress_data, progress_bar, progress_label, root)
root.mainloop()
