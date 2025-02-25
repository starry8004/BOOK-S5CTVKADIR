import csv
import os
import glob
from datetime import datetime
from tkinter import filedialog, Tk
import logging
from openpyxl import Workbook
from collections import defaultdict

class SellingHoneyDataProcessor:
    def __init__(self):
        self.setup_logging()
        self.logger = logging.getLogger(__name__)
        self.keyword_data = defaultdict(list)

    def setup_logging(self):
        """로깅 설정"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler(f'sellinghoney_process_{datetime.now().strftime("%Y%m%d_%H%M")}.log')
            ]
        )

    def select_folder(self):
        """폴더 선택 다이얼로그"""
        root = Tk()
        root.withdraw()
        folder_path = filedialog.askdirectory(title="CSV 파일이 있는 폴더를 선택하세요")
        return folder_path

    def get_sheet_name(self, filename):
        """파일명에서 시트명 추출"""
        # "셀링하니_" 제거 및 날짜 부분 제거
        base_name = os.path.basename(filename)
        # 날짜 부분 제거 (2024-11-14.csv 형식)
        name_parts = base_name.split('_')[1:-1]  # 첫 부분과 날짜 부분 제외
        return '_'.join(name_parts)

    def process_csv_file(self, file_path):
        """CSV 파일 처리"""
        self.logger.info(f"처리 중: {os.path.basename(file_path)}")
        data_dict = defaultdict(list)
        
        with open(file_path, 'r', encoding='utf-8') as csvfile:
            csv_reader = csv.DictReader(csvfile)
            for row in csv_reader:
                keyword = row['키워드']  # 키워드 컬럼명 확인 필요
                data_dict[keyword].append(row)
        
        # 중복 데이터 처리
        processed_data = []
        for keyword, rows in data_dict.items():
            if len(rows) > 1:
                # 중복된 경우 카테고리 데이터 합치기
                merged_row = rows[0].copy()
                categories = set()
                for row in rows:
                    if '카테고리' in row:  # 카테고리 컬럼명 확인 필요
                        categories.add(row['카테고리'])
                merged_row['카테고리'] = '|'.join(categories)
                processed_data.append(merged_row)
            else:
                processed_data.append(rows[0])
        
        # 검색량 기준 정렬
        return sorted(processed_data, 
                     key=lambda x: int(x.get('검색량', '0').replace(',', '')), 
                     reverse=True)

    def write_to_excel(self, all_data, output_path):
        """Excel 파일 생성"""
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거

        for sheet_name, data in all_data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # 헤더 쓰기
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 데이터 쓰기
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, key in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))

        wb.save(output_path)
        self.logger.info(f"파일 저장 완료: {output_path}")

    def process_files(self):
        """전체 처리 프로세스"""
        folder_path = self.select_folder()
        if not folder_path:
            self.logger.error("폴더가 선택되지 않았습니다.")
            return

        csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
        if not csv_files:
            self.logger.error("CSV 파일을 찾을 수 없습니다.")
            return

        # 데이터 처리
        processed_data = {}
        for file_path in csv_files:
            sheet_name = self.get_sheet_name(file_path)
            processed_data[sheet_name] = self.process_csv_file(file_path)

        # 결과 파일 저장
        output_filename = f"셀링하니_통합_키워드_데이터_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        output_path = os.path.join(folder_path, output_filename)
        self.write_to_excel(processed_data, output_path)

if __name__ == "__main__":
    processor = SellingHoneyDataProcessor()
    processor.process_files()