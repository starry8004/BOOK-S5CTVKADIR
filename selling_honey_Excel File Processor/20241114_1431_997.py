import csv
import os
from tkinter import filedialog
import tkinter as tk
from collections import defaultdict
import openpyxl
from datetime import datetime

def select_directory():
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="CSV 파일이 있는 폴더를 선택하세요")
    return folder_path

def read_csv_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            # 메모리 효율성을 위해 필요한 컬럼만 읽기
            needed_fields = {'키워드', '카테고리전체', '검색량', '경쟁률', '광고경쟁강도', '계절성'}
            # 순서 컬럼 찾기
            order_field = next((field for field in reader.fieldnames if '순서' in field or '검색량순' in field), None)
            if order_field:
                needed_fields.add(order_field)
            
            # 필요한 필드만 포함하는 새로운 딕셔너리 생성
            return [{k: row[k] for k in needed_fields if k in row} for row in reader]
    except Exception as e:
        print(f"파일 읽기 오류: {e}")
        return []

def process_data(data):
    if not data:
        return []

    # 키워드별 데이터 정리를 위한 딕셔너리
    keyword_data = {}
    
    # 한 번의 순회로 데이터 처리
    for row in data:
        keyword = row.get('키워드')
        if not keyword:
            continue
            
        if keyword not in keyword_data:
            # 첫 번째 등장하는 데이터 사용
            keyword_data[keyword] = row.copy()
            keyword_data[keyword]['카테고리전체'] = {row.get('카테고리전체', '')}
        else:
            # 카테고리 정보만 추가
            keyword_data[keyword]['카테고리전체'].add(row.get('카테고리전체', ''))
    
    # 카테고리 세트를 문자열로 변환
    for data in keyword_data.values():
        data['카테고리전체'] = ' | '.join(filter(None, data['카테고리전체']))
    
    # 리스트로 변환
    processed_data = list(keyword_data.values())
    
    # 순서 컬럼 찾기
    order_column = next((key for key in processed_data[0].keys() if '순서' in key or '검색량순' in key), None)
    
    # 정렬
    if order_column:
        try:
            processed_data.sort(
                key=lambda x: int(x.get(order_column, '0')) if x.get(order_column, '0').isdigit() 
                else float('inf')
            )
        except Exception as e:
            print(f"정렬 중 오류 발생: {e}")
    
    return processed_data

def get_output_filename():
    return f"셀링하니_통합_키워드_데이터_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

def extract_sheet_name(filename):
    # 간단한 문자열 처리로 변경
    name = filename.replace('셀링하니_', '').replace('_2024-11-14.csv', '')
    return name[:31]  # Excel 시트명 제한

def save_to_excel(all_data, output_filename):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 시트 생성 및 데이터 쓰기 최적화
    for sheet_name, data in all_data.items():
        if not data:
            continue
            
        ws = wb.create_sheet(title=sheet_name)
        
        # 헤더 쓰기
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 데이터 쓰기 - 최적화된 방식
        for row_idx, row_data in enumerate(data, 2):
            row_values = [row_data.get(header, '') for header in headers]
            for col_idx, value in enumerate(row_values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    try:
        wb.save(output_filename)
    except Exception as e:
        print(f"파일 저장 중 오류 발생: {e}")
        alternative_filename = f"셀링하니_통합_키워드_데이터_alternative_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(alternative_filename)
        print(f"대체 파일명으로 저장됨: {alternative_filename}")

def process_csv_files():
    folder_path = select_directory()
    if not folder_path:
        print("폴더가 선택되지 않았습니다.")
        return

    all_processed_data = {}
    output_filename = get_output_filename()
    total_files = len([f for f in os.listdir(folder_path) if f.endswith('.csv')])
    processed_count = 0
    
    print(f"총 {total_files}개의 CSV 파일을 처리합니다...")
    
    for filename in os.listdir(folder_path):
        if not filename.endswith('.csv'):
            continue
            
        processed_count += 1
        print(f"\n[{processed_count}/{total_files}] {filename} 처리 중...")
        
        try:
            file_path = os.path.join(folder_path, filename)
            data = read_csv_file(file_path)
            processed_data = process_data(data)
            sheet_name = extract_sheet_name(filename)
            all_processed_data[sheet_name] = processed_data
            print(f"{filename} 처리 완료")
            
        except Exception as e:
            print(f"{filename} 처리 중 오류 발생: {e}")
            continue
    
    if all_processed_data:
        save_to_excel(all_processed_data, output_filename)
        print("\n모든 파일 처리가 완료되었습니다.")
        print(f"결과가 '{output_filename}' 파일에 저장되었습니다.")
    else:
        print("\n처리된 데이터가 없습니다.")

if __name__ == "__main__":
    process_csv_files()