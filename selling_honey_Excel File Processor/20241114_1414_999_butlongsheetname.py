import csv
import os
from tkinter import filedialog
import tkinter as tk
from collections import defaultdict
import openpyxl

def select_directory():
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="CSV 파일이 있는 폴더를 선택하세요")
    return folder_path

def read_csv_file(file_path):
    data = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        # 첫 번째 행의 컬럼명들을 출력하여 확인
        if reader.fieldnames:
            print("CSV 파일의 컬럼명:", reader.fieldnames)
        for row in reader:
            data.append(row)
    return data

def process_data(data):
    if not data:
        return []
        
    # 컬럼명 확인
    sample_row = data[0]
    print("데이터 샘플의 키:", list(sample_row.keys()))
    
    # 순서를 나타내는 컬럼명 찾기
    order_column = None
    for key in sample_row.keys():
        if '순서' in key or '검색량순' in key:
            order_column = key
            print(f"순서 컬럼명 찾음: {order_column}")
            break
    
    if not order_column:
        print("순서 관련 컬럼을 찾을 수 없습니다.")
        return data

    # 키워드를 기준으로 데이터 정리
    keyword_data = defaultdict(list)
    keyword_column = '키워드'  # 키워드 컬럼명이 다르다면 이 부분도 수정 필요
    category_column = '카테고리전체'  # 카테고리 컬럼명이 다르다면 이 부분도 수정 필요
    
    for row in data:
        if keyword_column in row:
            keyword = row[keyword_column]
            keyword_data[keyword].append(row)
    
    # 중복 제거하고 카테고리 합치기
    processed_data = []
    for keyword, rows in keyword_data.items():
        base_row = rows[0].copy()
        
        if category_column in base_row:
            categories = [row[category_column] for row in rows if category_column in row]
            base_row[category_column] = ' | '.join(set(categories))
        
        processed_data.append(base_row)
    
    # 순서로 정렬
    try:
        processed_data.sort(
            key=lambda x: int(x[order_column]) if x[order_column].isdigit() 
            else float('inf')
        )
    except (KeyError, ValueError) as e:
        print(f"정렬 중 오류 발생: {e}")
        # 정렬에 실패하더라도 처리된 데이터는 반환
    
    return processed_data

def save_to_excel(all_data, output_filename='통합_키워드_데이터.xlsx'):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    for filename, data in all_data.items():
        sheet_name = os.path.splitext(filename)[0]
        # Excel 시트 이름 제한 (31자)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data[header])
    
    wb.save(output_filename)

def process_csv_files():
    folder_path = select_directory()
    if not folder_path:
        print("폴더가 선택되지 않았습니다.")
        return

    all_processed_data = {}
    
    for filename in os.listdir(folder_path):
        if filename.endswith('.csv'):
            print(f"\n{filename} 처리 중...")
            file_path = os.path.join(folder_path, filename)
            
            try:
                data = read_csv_file(file_path)
                processed_data = process_data(data)
                all_processed_data[filename] = processed_data
                print(f"{filename} 처리 완료")
            except Exception as e:
                print(f"{filename} 처리 중 오류 발생: {e}")
                continue
    
    if all_processed_data:
        save_to_excel(all_processed_data)
        print("\n모든 파일 처리가 완료되었습니다.")
        print("결과가 '통합_키워드_데이터.xlsx' 파일에 저장되었습니다.")
    else:
        print("\n처리된 데이터가 없습니다.")

if __name__ == "__main__":
    process_csv_files()