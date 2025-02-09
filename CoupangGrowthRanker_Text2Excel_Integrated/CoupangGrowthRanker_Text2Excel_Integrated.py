import tkinter as tk
from tkinter import filedialog
import openpyxl
import re
from tqdm import tqdm
import time
import os
from datetime import datetime
from openpyxl import Workbook

# 엑셀 파일 선택 (CoupangCategoryGrowthRanker 용)
def select_excel_file():
    """Tkinter를 사용하여 엑셀 파일 선택 다이얼로그를 열고, 선택된 파일 경로를 반환합니다."""
    root = tk.Tk()
    root.withdraw()  # Tk 창 숨김
    file_path = filedialog.askopenfilename(
        title="엑셀 파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return file_path

# 엑셀 파일 불러오기
def load_workbook_file(file_path):
    """openpyxl을 사용하여 엑셀 파일을 불러오고, 기본 워크시트를 반환합니다."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    return ws

# 엑셀 파일의 첫 행을 헤더로 취급
def get_header(ws):
    """첫 번째 행을 헤더로 간주하여 각 셀의 값을 리스트로 반환합니다."""
    header = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        header.append(cell.value)
    return header

# 날짜 형식 컬럼(YYYY-MM-DD)을 찾아 인덱스 반환
def extract_date_columns_indices(header):
    """
    헤더 리스트 중 'YYYY-MM-DD' 형태인 날짜 컬럼들을 찾아내고,
    날짜 순(내림차순: 최신 → 과거)으로 정렬한 후 인덱스 리스트를 반환합니다.
    """
    date_cols = []
    for idx, col in enumerate(header):
        if isinstance(col, str) and re.match(r'\d{4}-\d{2}-\d{2}', col):
            try:
                date_obj = datetime.strptime(col, "%Y-%m-%d")
                date_cols.append((idx, date_obj))
            except Exception:
                pass
    # 최신 날짜 순(내림차순)으로 정렬
    date_cols.sort(key=lambda x: x[1], reverse=True)
    sorted_indices = [idx for idx, _ in date_cols]
    return sorted_indices

# 카테고리 관련 컬럼 인덱스 추출
def extract_category_indices(header):
    """
    헤더에서 카테고리 관련 컬럼("대카테고리", "중카테고리", "소카테고리", "세부카테고리")의 인덱스를 추출합니다.
    """
    category_indices = {}
    for idx, col in enumerate(header):
        if col in ["대카테고리", "중카테고리", "소카테고리", "세부카테고리"]:
            category_indices[col] = idx
    return category_indices

# 각 행에 대해 총합, 최근 데이터 합계, 성장률, 최종 점수 계산
def process_rows(ws, header, date_indices, category_indices, recent_count=3):
    """
    각 행에 대해:
      - 전체 날짜 데이터 합계(total)
      - 최신 recent_count 개 날짜의 합계(recent_sum)
      - 이전 합계(older_sum)
      - 성장률(growth_rate)
      - 최종 점수(final_score = total * (1 + growth_rate))
    를 계산하고, 카테고리 정보와 함께 딕셔너리 형태로 리스트에 저장합니다.
    진행 상황은 tqdm로 표시합니다.
    """
    rows_data = []
    total_rows = ws.max_row - 1  # 헤더 제외
    for row in tqdm(ws.iter_rows(min_row=2, max_row=ws.max_row), total=total_rows, desc="Processing rows"):
        row_values = [cell.value for cell in row]
        # 전체 합계 계산
        total = 0
        for idx in date_indices:
            try:
                val = row_values[idx]
                if isinstance(val, (int, float)):
                    total += val
            except Exception:
                pass
        # 최신 날짜 데이터 합계 계산 (recent_count 개)
        recent_sum = 0
        for idx in date_indices[:recent_count]:
            try:
                val = row_values[idx]
                if isinstance(val, (int, float)):
                    recent_sum += val
            except Exception:
                pass
        older_sum = total - recent_sum
        if older_sum > 0:
            growth_rate = recent_sum / older_sum
        else:
            growth_rate = recent_sum  # 이전 합계가 0인 경우
        
        # 최종 점수 계산 (예: 총합과 성장률을 함께 고려)
        final_score = total * (1 + growth_rate)
        
        # 카테고리 정보 추출
        data = {}
        for key, idx in category_indices.items():
            data[key] = row_values[idx]
        data['total'] = total
        data['recent_sum'] = recent_sum
        data['older_sum'] = older_sum
        data['growth_rate'] = growth_rate
        data['final_score'] = final_score
        rows_data.append(data)
    return rows_data

# 성장률 기준으로 내림차순 정렬 후 순위 부여
def rank_categories_by_growth(rows_data):
    """
    전체 카테고리를 성장률(growth_rate)을 기준으로 내림차순 정렬한 후,
    순위를 1위부터 부여합니다.
    """
    rows_data_sorted = sorted(rows_data, key=lambda x: x['growth_rate'], reverse=True)
    for rank, data in enumerate(rows_data_sorted, start=1):
        data['rank'] = rank
    return rows_data_sorted

# 계산된 결과를 엑셀 파일로 바로 저장하는 함수
def save_results_to_excel(ranked_data, original_file_path):
    """
    계산된 카테고리 성장률 결과를 엑셀 파일로 저장합니다.
    파일명은 원본 파일명을 기반으로 현재 날짜와 시간을 포함합니다.
    """
    base_name = os.path.splitext(os.path.basename(original_file_path))[0]
    now_str = datetime.now().strftime("%Y-%m-%d_%H%M")
    output_file = f"{base_name}_{now_str}_growthRanked.xlsx"
    
    wb = Workbook()
    ws = wb.active
    
    # 저장할 컬럼 정의
    columns = ["rank", "대카테고리", "중카테고리", "소카테고리", "세부카테고리", 
               "total", "recent_sum", "older_sum", "growth_rate", "final_score"]
    ws.append(columns)
    
    for data in ranked_data:
        row = [data.get(col, "") for col in columns]
        ws.append(row)
    
    wb.save(output_file)
    print(f"\n엑셀 파일 저장 완료: {output_file}")

# 메인 함수: 데이터를 계산하고, 엑셀 파일로 저장하는 전체 과정을 수행
def main():
    file_path = select_excel_file()
    if not file_path:
        print("파일이 선택되지 않았습니다.")
        return
    
    print("엑셀 파일을 불러오는 중입니다...")
    ws = load_workbook_file(file_path)
    
    header = get_header(ws)
    print("헤더:", header)
    
    date_indices = extract_date_columns_indices(header)
    if not date_indices:
        print("날짜 형식의 컬럼이 발견되지 않았습니다.")
        return
    print("날짜 컬럼 인덱스:", date_indices)
    
    category_indices = extract_category_indices(header)
    if not category_indices:
        print("카테고리 관련 컬럼이 발견되지 않았습니다.")
        return
    print("카테고리 컬럼 인덱스:", category_indices)
    
    print("각 행의 데이터를 계산 중 (전체 합계, 최근 합계, 성장률 등)...")
    rows_data = process_rows(ws, header, date_indices, category_indices, recent_count=3)
    time.sleep(1)
    
    ranked_data = rank_categories_by_growth(rows_data)
    
    print("\n전체 카테고리 성장률 순위:")
    for data in ranked_data:
        print(f"Rank {data['rank']}: {data}")
    
    # 텍스트 파일 생성 없이 바로 엑셀 파일로 저장
    save_results_to_excel(ranked_data, file_path)

if __name__ == "__main__":
    main()
