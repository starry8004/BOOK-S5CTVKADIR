import tkinter as tk
from tkinter import filedialog
import openpyxl
import re
from tqdm import tqdm
import time
import os
from datetime import datetime

def select_excel_file():
    """
    Tkinter를 사용하여 엑셀 파일 선택 다이얼로그를 열고,
    선택된 파일 경로를 반환합니다.
    """
    root = tk.Tk()
    root.withdraw()  # Tk 창 숨김
    file_path = filedialog.askopenfilename(
        title="엑셀 파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return file_path

def load_workbook_file(file_path):
    """
    openpyxl을 사용해 엑셀 파일을 불러오고, 기본 워크시트를 반환합니다.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    return ws

def get_header(ws):
    """
    첫 번째 행을 헤더로 간주하여, 각 셀의 값을 리스트로 반환합니다.
    """
    header = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        header.append(cell.value)
    return header

def extract_date_columns_indices(header):
    """
    헤더 리스트 중 'YYYY-MM-DD' 형태인 날짜 컬럼들을 찾아내고,
    날짜 순(내림차순: 최신 → 과거)으로 정렬한 후 인덱스 리스트를 반환합니다.
    """
    date_cols = []
    for idx, col in enumerate(header):
        if isinstance(col, str) and re.match(r'\d{4}-\d{2}-\d{2}', col):
            try:
                date_obj = datetime.strptime(col, "%Y-%m-%d")
                date_cols.append((idx, date_obj))
            except Exception as e:
                pass
    # 최신 날짜 순(내림차순)으로 정렬
    date_cols.sort(key=lambda x: x[1], reverse=True)
    sorted_indices = [idx for idx, date_obj in date_cols]
    return sorted_indices

def extract_category_indices(header):
    """
    헤더에서 카테고리 관련 컬럼("대카테고리", "중카테고리", "소카테고리", "세부카테고리")의 인덱스를 추출합니다.
    """
    category_indices = {}
    for idx, col in enumerate(header):
        if col in ["대카테고리", "중카테고리", "소카테고리", "세부카테고리"]:
            category_indices[col] = idx
    return category_indices

def process_rows(ws, header, date_indices, category_indices, recent_count=3):
    """
    각 행에 대해:
      - 전체 날짜 데이터 합계(total)
      - 최신 recent_count 개 날짜의 합계(recent_sum)
      - 이전 합계(older_sum)
      - 성장률(growth_rate)
      - 최종 점수(final_score = total * (1 + growth_rate))
    를 계산하고, 카테고리 정보와 함께 딕셔너리 형태로 리스트에 저장합니다.
    진행 상황은 tqdm로 표시합니다.
    """
    rows_data = []
    total_rows = ws.max_row - 1  # 헤더 제외
    for row in tqdm(ws.iter_rows(min_row=2, max_row=ws.max_row), total=total_rows, desc="Processing rows"):
        row_values = [cell.value for cell in row]
        # 전체 합계 계산
        total = 0
        for idx in date_indices:
            try:
                val = row_values[idx]
                if isinstance(val, (int, float)):
                    total += val
            except Exception as e:
                pass
        # 최신 날짜 데이터 합계 계산 (recent_count개)
        recent_sum = 0
        for idx in date_indices[:recent_count]:
            try:
                val = row_values[idx]
                if isinstance(val, (int, float)):
                    recent_sum += val
            except Exception as e:
                pass
        older_sum = total - recent_sum
        if older_sum > 0:
            growth_rate = recent_sum / older_sum
        else:
            growth_rate = recent_sum  # 이전 합계가 0일 경우
        
        # 최종 점수 계산: 전체 합계에 성장률을 가중치로 반영
        final_score = total * (1 + growth_rate)
        
        # 카테고리 정보 추출
        data = {}
        for key, idx in category_indices.items():
            data[key] = row_values[idx]
        data['total'] = total
        data['recent_sum'] = recent_sum
        data['older_sum'] = older_sum
        data['growth_rate'] = growth_rate
        data['final_score'] = final_score
        rows_data.append(data)
    return rows_data

def rank_categories(rows_data, top_n=31):
    """
    최종 점수(final_score)를 기준으로 내림차순 정렬한 후,
    상위 top_n(예시에서는 31)개의 카테고리를 추출하고,
    순위를 1위부터 부여합니다.
    """
    rows_data_sorted = sorted(rows_data, key=lambda x: x['final_score'], reverse=True)
    top_categories = rows_data_sorted[:top_n]
    for rank, data in enumerate(top_categories, start=1):
        data['rank'] = rank
    return top_categories

def save_results(top_categories, original_file_path):
    """
    추천 결과(1위부터 31위까지)를 원본 파일명에 현재 날짜와 시간, 그리고 '_추천카테고리'를 붙여 텍스트 파일로 저장합니다.
    """
    base_name = os.path.splitext(os.path.basename(original_file_path))[0]
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    new_filename = f"{base_name}_{timestamp}_추천카테고리.txt"
    
    with open(new_filename, "w", encoding="utf-8") as f:
        f.write("추천 카테고리 순위 (Top 31):\n")
        f.write("====================================\n")
        for data in top_categories:
            f.write(f"Rank {data['rank']}: {data}\n")
    
    print(f"\n결과가 '{new_filename}' 파일에 저장되었습니다.")

def main():
    file_path = select_excel_file()
    if not file_path:
        print("파일이 선택되지 않았습니다.")
        return

    print("엑셀 파일을 불러오는 중입니다...")
    ws = load_workbook_file(file_path)
    
    header = get_header(ws)
    print("헤더:", header)
    
    date_indices = extract_date_columns_indices(header)
    if not date_indices:
        print("날짜 형식의 컬럼이 발견되지 않았습니다.")
        return
    print("날짜 컬럼 인덱스:", date_indices)
    
    category_indices = extract_category_indices(header)
    if not category_indices:
        print("카테고리 관련 컬럼이 발견되지 않았습니다.")
        return
    print("카테고리 컬럼 인덱스:", category_indices)
    
    print("각 행의 데이터를 계산 중 (총합, 최근 합계, 성장률, 최종 점수)...")
    rows_data = process_rows(ws, header, date_indices, category_indices, recent_count=3)
    time.sleep(1)
    
    top_categories = rank_categories(rows_data, top_n=31)
    print("\n추천 카테고리 순위 (Top 31):")
    for data in top_categories:
        print(f"Rank {data['rank']}: {data}")
    
    save_results(top_categories, file_path)
    
if __name__ == "__main__":
    main()
