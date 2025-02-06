import tkinter as tk
from tkinter import filedialog
import openpyxl
import re
from tqdm import tqdm
import time

def select_excel_file():
    """
    Tkinter를 사용하여 엑셀 파일 선택 다이얼로그를 연 후,
    선택된 파일 경로를 반환합니다.
    """
    root = tk.Tk()
    root.withdraw()  # Tk 창 숨김
    file_path = filedialog.askopenfilename(
        title="엑셀 파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return file_path

def load_workbook_file(file_path):
    """
    openpyxl을 사용하여 엑셀 파일을 불러오고, 기본 워크시트를 반환합니다.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    return ws

def get_header(ws):
    """
    첫 번째 행을 헤더로 간주하여, 각 셀의 값을 리스트로 반환합니다.
    """
    header = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        header.append(cell.value)
    return header

def is_date_column(value):
    """
    컬럼명이 'YYYY-MM-DD' 형태인지 확인합니다.
    """
    if isinstance(value, str) and re.match(r'\d{4}-\d{2}-\d{2}', value):
        return True
    return False

def extract_date_columns_indices(header):
    """
    헤더 리스트에서 날짜 형식(예: "2025-02-03")에 해당하는 컬럼 인덱스만 추출합니다.
    """
    date_cols = []
    for idx, col in enumerate(header):
        if is_date_column(col):
            date_cols.append(idx)
    return date_cols

def extract_category_indices(header):
    """
    헤더에서 카테고리 관련 컬럼("대카테고리", "중카테고리", "소카테고리", "세부카테고리")의 인덱스를 추출합니다.
    """
    category_indices = {}
    for idx, col in enumerate(header):
        if col in ["대카테고리", "중카테고리", "소카테고리", "세부카테고리"]:
            category_indices[col] = idx
    return category_indices

def process_rows(ws, header, date_indices, category_indices):
    """
    각 행에 대해 날짜 형식 컬럼의 값을 합산하고, 
    카테고리 정보와 함께 딕셔너리 형태로 리스트에 저장합니다.
    tqdm를 사용하여 진행 상황을 표시합니다.
    """
    rows_data = []
    total_rows = ws.max_row - 1  # 헤더를 제외한 행의 수
    for row in tqdm(ws.iter_rows(min_row=2, max_row=ws.max_row), total=total_rows, desc="Processing rows"):
        row_values = [cell.value for cell in row]
        total = 0
        for idx in date_indices:
            try:
                val = row_values[idx]
                if isinstance(val, (int, float)):
                    total += val
            except Exception as e:
                pass  # 값이 없거나 숫자가 아닌 경우 건너뜀
        # 카테고리 정보 추출
        data = {}
        for key, idx in category_indices.items():
            data[key] = row_values[idx]
        data['total'] = total
        rows_data.append(data)
    return rows_data

def recommend_categories(rows_data, top_n=5):
    """
    'total' 값을 기준으로 내림차순 정렬 후 상위 top_n개의 카테고리를 반환합니다.
    """
    rows_data_sorted = sorted(rows_data, key=lambda x: x['total'], reverse=True)
    return rows_data_sorted[:top_n]

def main():
    file_path = select_excel_file()
    if not file_path:
        print("파일이 선택되지 않았습니다.")
        return

    print("엑셀 파일을 불러오는 중입니다...")
    ws = load_workbook_file(file_path)
    
    header = get_header(ws)
    print("헤더:", header)
    
    date_indices = extract_date_columns_indices(header)
    if not date_indices:
        print("날짜 형식의 컬럼이 발견되지 않았습니다.")
        return
    print("날짜 컬럼 인덱스:", date_indices)
    
    category_indices = extract_category_indices(header)
    if not category_indices:
        print("카테고리 관련 컬럼이 발견되지 않았습니다.")
        return
    print("카테고리 컬럼 인덱스:", category_indices)
    
    print("각 행의 날짜 데이터 합계를 계산하는 중...")
    rows_data = process_rows(ws, header, date_indices, category_indices)
    time.sleep(1)
    
    recommended = recommend_categories(rows_data, top_n=5)
    print("\n추천 카테고리 (Top 5):")
    for data in recommended:
        print(data)
    
if __name__ == "__main__":
    main()
